import csv
import inspect
import logging
import xlrd
from openpyxl import workbook, load_workbook


class util():
    def __init__(self,Driver):
        self.driver=Driver
    def forloop(self,list):
        for stop in list:
            i = 1
            if (stop.text == "1 Stop"):
                print("Test case", i, "PASSED as having", stop.text)
            else:
                print("Test case Fails at", i)
            print(stop.text)
            i + 1

    def custom_logger(logLevel=logging.DEBUG):
        # create logger and Set class/method name from where it called
        logger_name= inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        # create console hendler and file handler and set log level
        logger.setLevel(logLevel)
        fh = logging.FileHandler("yatra.log")
        # create formatter
        formatt = logging.Formatter('%(asctime)s - %(filename)s : %(message)s')
        # add formatter to the console of file handler
        fh.setFormatter(formatt)
        # add console handler to logger
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name,sheet):
        datalst=[]

        wb= load_workbook(filename=file_name)
        #wb=xlrd.open_workbook(filename=file_name)
        sh=wb[sheet]
        row_count=sh.max_row
        col_count= sh.column

        for i in range(2,row_count+1):
            rowdata=[]
            for j in range(1,col_count+1):
                rowdata.append(sh.cell(row=i, column=j).value)
            datalst.append(rowdata)

        return datalst

    def read_data_fromm_csv(file_name):
        csvdata_lst=[]
        #open Csv file
        csv_data= open(file_name,"r")
        #create csv reader
        csv_reader= csv.reader(csv_data)
        #to skip header
        next(csv_reader)
        #add data in list
        for r in csv_reader:
            csvdata_lst.append(r)
        return csvdata_lst